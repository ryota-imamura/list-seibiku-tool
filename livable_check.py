"""
リバブル重複チェックモジュール

作業リスト × リバブル一括リスト の突合を行い、
過去にどのセンターがいつ・どのリストで送ったかをExcelに出力する。

マッチキー: NFKC正規化(都道府県+市区町村) + NFKC正規化(オーナー名) ― 完全一致・切り捨てなし
"""

import io
import re
import unicodedata
from collections import defaultdict

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


# ── 正規化 ────────────────────────────────────────────────────────────

_PREF_RE = re.compile(r'^(東京都|北海道|(?:大阪|京都)府|.{2,3}県)')

def _nkfc(s):
    """全角→半角・前後スペース除去"""
    return unicodedata.normalize("NFKC", str(s)).strip() if pd.notna(s) else ""


def _city_prefix(address):
    """
    都道府県 + 市区町村 を抽出して返す。
    政令指定都市は「市+区」まで取得（川崎市高津区 / 川崎市多摩区 を区別）。
    """
    s = _nkfc(address)
    if not s:
        return ""
    m = _PREF_RE.match(s)
    if not m:
        return s[:8]
    pref = m.group(1)
    rest = s[len(pref):]
    # 政令指定都市: 「市」の直後に「区」が続くパターン
    ku_m = re.match(r'^([^\d一二三四五六七八九十百千万]+?市[^\d一二三四五六七八九十百千万]*?区)', rest)
    if ku_m:
        return pref + ku_m.group(1)
    # 通常の市区町村
    city_m = re.match(r'^([^\d一二三四五六七八九十百千万]+?[市区町村])', rest)
    if city_m:
        return pref + city_m.group(1)
    return pref + rest[:4]


def _build_key(address, owner_name):
    """
    マッチキー = NFKC正規化した（市区町村レベル住所 + オーナー名）
    ・10文字切り捨てなし → 別会社の誤マッチを防止
    ・NFKC正規化 → 全角/半角の表記ゆれを吸収（Ｓ→S など）
    """
    name = _nkfc(owner_name)
    if not name:
        return None
    prefix = _city_prefix(address)
    return prefix + name


# ── リバブル一括リスト インデックス構築 ─────────────────────────────────

def build_livable_index(livable_bytes):
    """
    リバブル一括リストをロードし、マッチキー → 送付履歴リスト のインデックスを返す。

    ※ リバブル一括リストでは「発注日」列にセンター名、「発送日」列に日付が入っている。
    """
    df = pd.read_excel(io.BytesIO(livable_bytes))

    # 「列3」（truncated key）は使わず、住所+名前から新規計算
    index = defaultdict(list)
    for _, row in df.iterrows():
        key = _build_key(row.get("オーナー住所", ""), row.get("オーナー名", ""))
        if not key:
            continue
        center = _nkfc(row.get("発注日", ""))   # 発注日列にセンター名が入っている
        list_name = _nkfc(row.get("リスト名", ""))
        send_date = _nkfc(row.get("発送日", ""))
        index[key].append({
            "リスト名": list_name,
            "センター名": center,
            "発送日": send_date,
        })

    return dict(index)


# ── メイン処理 ────────────────────────────────────────────────────────

def check_duplicates(
    work_bytes,
    livable_bytes,
    work_sheet=0,
    progress_callback=None,
):
    """
    作業リスト × リバブル一括リスト の重複チェック。

    Returns:
        (excel_bytes, summary_dict)
    """

    def notify(msg, pct=None):
        if progress_callback:
            progress_callback(msg, pct)

    notify("リバブル一括リストを読み込み中...", 0.05)
    livable_index = build_livable_index(livable_bytes)
    notify(f"インデックス構築完了: {len(livable_index):,}件のユニークキー", 0.20)

    df_work = pd.read_excel(io.BytesIO(work_bytes), sheet_name=work_sheet)
    total = len(df_work)
    notify(f"作業リスト読み込み完了: {total}行", 0.25)

    match_results = []
    for i, row in df_work.iterrows():
        key = _build_key(row.get("オーナー住所", ""), row.get("オーナー名", ""))
        matches = livable_index.get(key, []) if key else []
        match_results.append(matches)
        if (i + 1) % 50 == 0 or i + 1 == total:
            notify(f"マッチング中... ({i+1}/{total})", 0.25 + 0.55 * (i + 1) / total)

    # ── 付加列を生成 ──
    dup_flags, match_counts = [], []
    latest_centers, latest_lists, latest_dates, past_lists = [], [], [], []

    for matches in match_results:
        if matches:
            # リスト名単位で集約（同一リストの複数行は1件に）
            seen = {}
            for m in matches:
                ln = m["リスト名"]
                if ln not in seen:
                    seen[ln] = m
            deduped = sorted(seen.values(), key=lambda x: x["発送日"], reverse=True)
            # 最新1件 + 過去分
            latest = deduped[0]
            older  = deduped[1:]

            dup_flags.append("有")
            match_counts.append(len(deduped))
            latest_centers.append(latest["センター名"])
            latest_lists.append(latest["リスト名"])
            latest_dates.append(latest["発送日"])
            past_lists.append(" / ".join(d["リスト名"] for d in older) if older else "")
        else:
            dup_flags.append("無")
            match_counts.append(0)
            latest_centers.append("")
            latest_lists.append("")
            latest_dates.append("")
            past_lists.append("")

    df_out = df_work.copy()
    df_out["リバブル重複"]    = dup_flags
    df_out["重複件数"]        = match_counts
    df_out["最新センター名"]  = latest_centers
    df_out["最新リスト名"]    = latest_lists
    df_out["最新発送日"]      = latest_dates
    df_out["過去リスト名"]    = past_lists

    notify("Excel出力を作成中...", 0.85)

    added_cols = ["リバブル重複", "重複件数", "最新センター名", "最新リスト名", "最新発送日", "過去リスト名"]

    # ── 重複詳細（1マッチ1行）──
    detail_rows = []
    for row, matches in zip(df_work.itertuples(index=False), match_results):
        if not matches:
            continue
        seen = {}
        for m in matches:
            if m["リスト名"] not in seen:
                seen[m["リスト名"]] = m
        for m in sorted(seen.values(), key=lambda x: x["発送日"], reverse=True):
            detail_rows.append({
                "オーナー名":  getattr(row, "オーナー名", ""),
                "郵便番号":    getattr(row, "郵便番号", ""),
                "オーナー住所": getattr(row, "オーナー住所", ""),
                "センター名":  m["センター名"],
                "リスト名":    m["リスト名"],
                "発送日":      m["発送日"],
            })

    df_detail = pd.DataFrame(detail_rows) if detail_rows else pd.DataFrame(
        columns=["オーナー名", "郵便番号", "オーナー住所", "センター名", "リスト名", "発送日"]
    )

    # ── センター別集計 ──
    center_stats = defaultdict(int)
    for r in detail_rows:
        center_stats[r["センター名"]] += 1

    # ── Excel作成 ──
    wb = openpyxl.Workbook()
    H_COLOR = {"blue": "1F4E79", "red": "C00000", "purple": "7030A0", "green": "375623"}

    def style_header(cell, color_key):
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill = PatternFill("solid", start_color=H_COLOR[color_key])
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def write_df(ws, df, color_key, col_widths=None):
        for ci, col in enumerate(df.columns, 1):
            style_header(ws.cell(row=1, column=ci, value=col), color_key)
        for ri, row in enumerate(df.itertuples(index=False), 2):
            for ci, val in enumerate(row, 1):
                ws.cell(row=ri, column=ci, value="" if pd.isna(val) else val)
        if col_widths:
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.row_dimensions[1].height = 20

    # シート1: 重複チェック済み作業リスト
    ws1 = wb.active
    ws1.title = "重複チェック済みリスト"
    base_cols = list(df_work.columns)
    all_cols = base_cols + added_cols
    for ci, col in enumerate(all_cols, 1):
        style_header(ws1.cell(row=1, column=ci, value=col), "blue")
    FILL_DUP = PatternFill("solid", start_color="FCE4D6")
    for ri, (row_vals, flag) in enumerate(
        zip(df_out[all_cols].itertuples(index=False), dup_flags), 2
    ):
        for ci, val in enumerate(row_vals, 1):
            cell = ws1.cell(row=ri, column=ci, value="" if pd.isna(val) else val)
            if flag == "有":
                cell.fill = FILL_DUP
    base_widths = [max(12, len(str(c)) + 2) for c in base_cols]
    added_widths = [10, 8, 35, 55, 12, 55]
    for i, w in enumerate(base_widths + added_widths, 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws1.row_dimensions[1].height = 20

    # シート2: 重複詳細（1マッチ1行）
    ws2 = wb.create_sheet("重複詳細")
    write_df(ws2, df_detail, "red", col_widths=[18, 13, 40, 35, 60, 15])

    # シート3: センター別集計
    ws3 = wb.create_sheet("センター別集計")
    center_df = pd.DataFrame(
        sorted(center_stats.items(), key=lambda x: -x[1]),
        columns=["センター名", "重複件数"],
    ) if center_stats else pd.DataFrame(columns=["センター名", "重複件数"])
    write_df(ws3, center_df, "purple", col_widths=[40, 12])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    dup_count = sum(1 for f in dup_flags if f == "有")
    summary = {
        "total": total,
        "dup_count": dup_count,
        "clean_count": total - dup_count,
        "center_stats": dict(center_stats),
    }
    notify("完了", 1.0)
    return buf.read(), summary


# ── CLI 実行 ──────────────────────────────────────────────────────────

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="リバブル重複チェックツール")
    parser.add_argument("work_list",    help="作業リスト Excelファイルパス")
    parser.add_argument("livable_list", help="リバブル一括リスト Excelファイルパス")
    parser.add_argument("-o", "--output", default="重複チェック結果.xlsx")
    parser.add_argument("-s", "--sheet",  default=0)
    args = parser.parse_args()

    sheet = int(args.sheet) if str(args.sheet).isdigit() else args.sheet
    with open(args.work_list,    "rb") as f: work_bytes    = f.read()
    with open(args.livable_list, "rb") as f: livable_bytes = f.read()

    def progress(msg, pct):
        bar = int((pct or 0) * 30)
        print(f"\r[{'#'*bar}{'.'*(30-bar)}] {msg}      ", end="", flush=True)

    excel_bytes, summary = check_duplicates(work_bytes, livable_bytes, sheet, progress)
    print()
    with open(args.output, "wb") as f:
        f.write(excel_bytes)

    print(f"\n=== 結果サマリー ===")
    print(f"  総行数   : {summary['total']:,}")
    print(f"  重複あり : {summary['dup_count']:,}")
    print(f"  重複なし : {summary['clean_count']:,}")
    for center, cnt in sorted(summary["center_stats"].items(), key=lambda x: -x[1]):
        print(f"    {center}: {cnt}件")
    print(f"\n出力: {args.output}")
