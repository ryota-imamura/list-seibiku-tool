import io
import os
import re
import time
import json
import sqlite3
import datetime
import unicodedata
import urllib.request
import urllib.parse

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def _get_app_version():
    """現在のGitコミットSHAを返す。Streamlit Cloudでも動作"""
    try:
        import subprocess
        cwd = os.path.dirname(os.path.abspath(__file__))
        return subprocess.check_output(
            ['git', 'log', '-1', '--format=%h (%cd)', '--date=format:%Y-%m-%d %H:%M'],
            cwd=cwd, stderr=subprocess.DEVNULL, timeout=2).decode().strip()
    except Exception:
        return 'unknown'

# ── posuto (日本郵便公式CSV内包DB) ──────────────────────────────────
# posuto は約80MBのSQLite DBで、住所→郵便番号の完全な逆引きを可能にする
_POSUTO_CONN = None

def _get_posuto_conn():
    global _POSUTO_CONN
    if _POSUTO_CONN is not None:
        return _POSUTO_CONN
    try:
        import posuto
        db_path = os.path.join(os.path.dirname(posuto.__file__), 'postaldata.db')
        _POSUTO_CONN = sqlite3.connect(db_path, check_same_thread=False)
        return _POSUTO_CONN
    except Exception:
        return None

# ── 正規化ユーティリティ ──────────────────────────────────────────────

def to_halfwidth(s):
    if not isinstance(s, str):
        return s
    return unicodedata.normalize('NFKC', s).strip()

def normalize_postal(s):
    if not isinstance(s, str):
        s = str(s) if pd.notna(s) else ""
    s = to_halfwidth(s)
    for ch in ["-", "ー", "−", "‐", "－"]:
        s = s.replace(ch, "")
    digits = re.sub(r'\D', '', s)
    if len(digits) == 7:
        return f"{digits[:3]}-{digits[3:]}"
    # Excelで数値型読込時の先頭ゼロ落ちを救済（6桁→0付与で7桁化）
    if len(digits) == 6:
        return f"0{digits[:2]}-{digits[2:]}"
    return None

def is_valid_postal(s):
    return bool(s and re.match(r'^\d{3}-\d{4}$', str(s)))

def is_garbled(s):
    if not isinstance(s, str):
        return False
    for p in [r'[\?]{3,}', r'[□]{3,}', r'[〓]{3,}', r'[■]{3,}', r'\?{3,}']:
        if re.search(p, s):
            return True
    return False

PREF_PATTERN = re.compile(r'^(東京都|北海道|(?:大阪|京都)府|.{2,3}県)')

def has_prefecture(address):
    return bool(address and PREF_PATTERN.match(address))

def normalize_address_for_compare(addr):
    if not addr:
        return ""
    s = to_halfwidth(addr)
    s = re.sub(r'(\d+)丁目', r'\1-', s)
    s = re.sub(r'(\d+)番地', r'\1-', s)
    s = re.sub(r'(\d+)番',  r'\1-', s)
    s = re.sub(r'(\d+)号',  r'\1',  s)
    s = re.sub(r'-+$', '', s)
    s = re.sub(r'-{2,}', '-', s)
    return s.strip()

# ── API呼び出し ───────────────────────────────────────────────────────

def _get_json(url, timeout=5):
    try:
        with urllib.request.urlopen(url, timeout=timeout) as r:
            return json.loads(r.read().decode('utf-8'))
    except Exception:
        return None

def lookup_address_from_postal(postal):
    code = postal.replace("-", "")
    data = _get_json(f"https://zipcloud.ibsnet.co.jp/api/search?zipcode={code}")
    if data and data.get('results'):
        res = data['results'][0]
        addr = res.get('address1','') + res.get('address2','') + res.get('address3','')
        return addr.strip() or None
    return None

def lookup_prefecture_from_postal(postal):
    if not is_valid_postal(postal):
        return None
    code = postal.replace("-", "")
    data = _get_json(f"https://zipcloud.ibsnet.co.jp/api/search?zipcode={code}")
    if data and data.get('results'):
        return data['results'][0].get('address1')
    return None

# 都道府県リスト
_ALL_PREFS = [
    '北海道','青森県','岩手県','宮城県','秋田県','山形県','福島県',
    '茨城県','栃木県','群馬県','埼玉県','千葉県','東京都','神奈川県',
    '新潟県','富山県','石川県','福井県','山梨県','長野県','岐阜県',
    '静岡県','愛知県','三重県','滋賀県','京都府','大阪府','兵庫県',
    '奈良県','和歌山県','鳥取県','島根県','岡山県','広島県','山口県',
    '徳島県','香川県','愛媛県','高知県','福岡県','佐賀県','長崎県',
    '熊本県','大分県','宮崎県','鹿児島県','沖縄県',
]
_CITY_PREF_CACHE = {}  # {city_name: pref}
_PREF_CITIES_CACHE = {}  # {pref: [city_name, ...]} 市町村すべて（政令市の区含む）
_TOWNS_CACHE = {}  # {(pref, city): [{'town': ..., 'postal': ...}, ...]}

# 廃止された旧市 → 現存する新市（政令市の区など）のマッピング
# 郡が付かないため郡フォールバックでは拾えない合併ケース
_OBSOLETE_CITY_MAP = {
    ('埼玉県', '浦和市'): ['さいたま市浦和区', 'さいたま市南区', 'さいたま市桜区', 'さいたま市緑区'],
    ('埼玉県', '大宮市'): ['さいたま市大宮区', 'さいたま市北区', 'さいたま市西区', 'さいたま市見沼区'],
    ('埼玉県', '与野市'): ['さいたま市中央区'],
    ('埼玉県', '岩槻市'): ['さいたま市岩槻区'],
    ('東京都', '秋川市'): ['あきる野市'],
    ('東京都', '田無市'): ['西東京市'],
    ('東京都', '保谷市'): ['西東京市'],
    ('静岡県', '清水市'): ['静岡市清水区'],
    ('静岡県', '蒲原町'): ['静岡市清水区'],
    ('福岡県', '若宮町'): ['宮若市'],
    ('熊本県', '富合町'): ['熊本市南区'],
    ('新潟県', '新津市'): ['新潟市秋葉区'],
    ('新潟県', '白根市'): ['新潟市南区'],
    ('新潟県', '豊栄市'): ['新潟市北区'],
    # 浜松市 2024年4月の区再編
    ('静岡県', '浜松市東区'): ['浜松市中央区'],
    ('静岡県', '浜松市西区'): ['浜松市中央区'],
    ('静岡県', '浜松市南区'): ['浜松市中央区'],
    ('静岡県', '浜松市北区'): ['浜松市中央区', '浜松市浜名区', '浜松市天竜区'],
    ('静岡県', '浜松市浜北区'): ['浜松市浜名区'],
}

# 漢字異体字の正規化テーブル（地名に頻出する異体字）
_KANJI_VARIANTS = str.maketrans({
    '螢': '蛍', '蛍': '蛍',
    '邊': '辺', '邉': '辺',
    '齋': '斎', '齊': '斉',
    '澤': '沢',
    '濱': '浜',
    '櫻': '桜',
    '應': '応',
    '會': '会',
    '舊': '旧',
    '寳': '宝', '寶': '宝',
    '靑': '青',
    '黑': '黒',
    '槇': '槙',
})

# 地名特有の表記揺れ（漢字 ↔ ひらがな・別表記）
_PLACE_NAME_VARIANTS = {
    '大埆': '大そね',  # 高知県南国市の地名
}

def _build_city_pref_cache():
    """全都道府県の市区町村→都道府県マッピングを構築（初回のみ）

    posuto DB（日本郵便公式CSV）から先に構築し、HeartRailsで上書き補完する。
    Streamlit Cloud等の環境で外部API（HeartRails）がブロック・タイムアウト
    した場合でも、posuto単独で確実にキャッシュが構築される。
    """
    global _CITY_PREF_CACHE, _PREF_CITIES_CACHE
    if _CITY_PREF_CACHE:
        return

    # まず posuto DB から市区町村→都道府県マッピングを構築（オフライン・確実）
    conn = _get_posuto_conn()
    if conn is not None:
        try:
            for pref, city_full in conn.execute(
                'SELECT DISTINCT prefecture, city FROM postal_data'):
                if not pref or not city_full:
                    continue
                _CITY_PREF_CACHE[city_full] = pref
                # 政令指定都市の親市も登録: 「北九州市小倉南区」→「北九州市」
                m = re.match(r'^(.+市)', city_full)
                if m:
                    _CITY_PREF_CACHE.setdefault(m.group(1), pref)
                    _PREF_CITIES_CACHE.setdefault(pref, [])
                    if m.group(1) not in _PREF_CITIES_CACHE[pref]:
                        _PREF_CITIES_CACHE[pref].append(m.group(1))
                _PREF_CITIES_CACHE.setdefault(pref, [])
                if city_full not in _PREF_CITIES_CACHE[pref]:
                    _PREF_CITIES_CACHE[pref].append(city_full)
        except Exception:
            pass

    # 続いて HeartRails で補完（呼び出し失敗時はposutoキャッシュで動作継続）
    # 既存の posuto キャッシュは上書きせず追記のみ
    for pref in _ALL_PREFS:
        url = f"https://geoapi.heartrails.com/api/json?method=getCities&prefecture={urllib.parse.quote(pref)}"
        data = _get_json(url)
        if not data:
            continue
        cities_in_pref = _PREF_CITIES_CACHE.setdefault(pref, [])
        for loc in data.get('response', {}).get('location', []):
            city_full = loc.get('city', '')
            if not city_full:
                continue
            _CITY_PREF_CACHE.setdefault(city_full, pref)
            # 政令指定都市: 「北九州市小倉南区」→「北九州市」も登録
            m = re.match(r'^(.+市)', city_full)
            if m:
                _CITY_PREF_CACHE.setdefault(m.group(1), pref)
                if m.group(1) not in cities_in_pref:
                    cities_in_pref.append(m.group(1))
            # 政令市の区・町・村も検索候補に含める（合併済み市町村の郵便番号逆引き用）
            if city_full not in cities_in_pref:
                cities_in_pref.append(city_full)
        time.sleep(0.3)

def _get_towns(pref, city):
    """HeartRails getTowns 結果をキャッシュ付きで取得"""
    key = (pref, city)
    if key in _TOWNS_CACHE:
        return _TOWNS_CACHE[key]
    url = (
        "https://geoapi.heartrails.com/api/json?method=getTowns"
        f"&prefecture={urllib.parse.quote(pref)}"
        f"&city={urllib.parse.quote(city)}"
    )
    data = _get_json(url)
    locations = []
    if data:
        locations = data.get('response', {}).get('location', []) or []
    _TOWNS_CACHE[key] = locations
    return locations

def lookup_prefecture_from_city(address):
    """住所の先頭から市区町村名を抽出して都道府県を逆引き"""
    _build_city_pref_cache()
    # 非欲張りで最短一致: 鳥栖市本鳥栖町→鳥栖市、北九州市小倉南区→北九州市
    for pat in [r'^(.{2,12}?[市区町村])', r'^(.{2,6}?[市区町村])']:
        m = re.match(pat, address)
        if m:
            city = m.group(1)
            if city in _CITY_PREF_CACHE:
                return _CITY_PREF_CACHE[city], city
    # 郡+旧町村の場合（合併済み）: 郡名の前部分+市 を試みる
    # 例: 神埼郡神埼町 → 神埼市
    gun_m = re.match(r'^(.+)郡', address)
    if gun_m:
        merged_city = gun_m.group(1) + '市'
        if merged_city in _CITY_PREF_CACHE:
            return _CITY_PREF_CACHE[merged_city], merged_city
    return None, None

def _extract_city_and_town(address):
    """住所から都道府県除去済み文字列の市区町村と町域を抽出"""
    # 「市/町/村/区」の出現位置を全て候補化し、_CITY_PREF_CACHE 一致を優先する
    # これにより「四日市市」「津市一志町」などの曖昧ケースを正しく解釈
    matches = [m for m in re.finditer(r'[市区町村]', address)]
    if not matches:
        return None, None
    candidates = []
    for m in matches:
        end = m.end()
        if end < 1 or end > 11:
            continue
        candidates.append(address[:end])
    if not candidates:
        return None, None
    # キャッシュに存在するものを優先（複数あればより長いものを優先）
    city = None
    cached = [c for c in candidates if c in _CITY_PREF_CACHE]
    if cached:
        city = max(cached, key=len)
    else:
        # キャッシュ未構築時 or 未登録の市: 最短候補を採用
        city = min(candidates, key=len)
    town_rest = address[len(city):]
    # 政令指定都市の区: city+区 がキャッシュ or 廃止旧区マップに存在する場合のみ拡張
    # （姫路市飾磨区など中核市の地名「区」を誤って取り込まない）
    if city.endswith('市'):
        ku_m = re.match(r'^([^\d一二三四五六七八九十]+区)', town_rest)
        if ku_m:
            candidate = city + ku_m.group(1)
            # キャッシュ または _OBSOLETE_CITY_MAP のキーに存在すれば区を含める
            if candidate in _CITY_PREF_CACHE or \
               any(k[1] == candidate for k in _OBSOLETE_CITY_MAP.keys()):
                city = candidate
                town_rest = town_rest[len(ku_m.group(1)):]
    # 先頭の「大字/字」を除去
    town_rest = re.sub(r'^[大小]?字', '', town_rest)
    # 最初に登場するアスキー数字（番地）以降を全て除去
    # 例: 「吹屋204-8」→「吹屋」、「福島799-7」→「福島」、「○○1丁目2番地」→「○○」
    town = re.sub(r'\d.*', '', town_rest).strip()
    # 漢数字+丁目/番地/号 のケースも除去（例: 「○○一丁目」）
    # 「番」は除外: 「九番町」などの固有名詞町名を残すため
    town = re.sub(r'[一二三四五六七八九十百千万]+(?:丁目|番地|号).*', '', town).strip()
    # 「○丁目」末尾（数字なしの漢数字丁目）も除去
    town = re.sub(r'[一二三四五六七八九十百千万]+丁目.*', '', town).strip()
    # 中間に含まれる「大字/小字」も除去（例: 大和町大字尼寺 → 大和町尼寺）
    town = re.sub(r'[大小]字', '', town).strip()
    return city, town

def _posuto_search(pref, city, town):
    """posuto DB で住所→郵便番号を逆引き

    戻り値: (postal | None, info_dict)
    info_dict は HeartRails 用と互換: {'status': 'ok'|'town_not_matched', 'candidates': [...]}
    """
    conn = _get_posuto_conn()
    if conn is None:
        return None, {'status': 'town_not_matched', 'candidates': []}
    if not town:
        return None, {'status': 'town_not_matched', 'candidates': []}

    def _norm(s):
        return s.replace('ヶ', 'ケ').replace('ヵ', 'カ').translate(_KANJI_VARIANTS)

    def _fmt(code):
        return f"{code[:3]}-{code[3:]}"

    town_n = _norm(town)
    town_strip = re.sub(r'[町村]$', '', town_n)

    # 1. 完全一致
    rows = conn.execute(
        'SELECT code, neighborhood FROM postal_data WHERE prefecture=? AND city=?',
        (pref, city or '')).fetchall()
    for code, nb in rows:
        if _norm(nb) == town_n:
            return _fmt(code), {'status': 'ok', 'candidates': []}
    # 2. 前方一致 (xx → xx町、xx → xx一丁目 など)
    for code, nb in rows:
        nb_n = _norm(nb)
        if nb_n.startswith(town_n) or (town_strip and nb_n.startswith(town_strip)):
            return _fmt(code), {'status': 'ok', 'candidates': []}
    # 3. 逆方向の前方一致 (鼻毛石 ← 鼻毛石町 などは1と2でカバー、町除き完全一致)
    for code, nb in rows:
        nb_n = _norm(nb)
        nb_strip = re.sub(r'[町村]$', '', nb_n)
        if town_strip and nb_strip == town_strip:
            return _fmt(code), {'status': 'ok', 'candidates': []}
    # 4. town の前方「○○町」または「○○村」までで切ったバリアントを試す
    #    例: 「広田町富田」→「広田町」、「吉村町図公甲」→「吉村町」
    # 「町」を優先、見つからなければ「村」を試す
    for end_ch in ('町', '村'):
        short_m = re.match(rf'^(.+?{end_ch})', town_n)
        if short_m and short_m.group(1) != town_n:
            short_town = short_m.group(1)
            for code, nb in rows:
                if _norm(nb) == short_town:
                    return _fmt(code), {'status': 'ok', 'candidates': []}
    # 5. 地名異体字テーブル（漢字→ひらがななど）を適用して再検索
    town_alt = town_n
    for k, v in _PLACE_NAME_VARIANTS.items():
        town_alt = town_alt.replace(k, v)
    if town_alt != town_n:
        for code, nb in rows:
            if _norm(nb).startswith(town_alt):
                return _fmt(code), {'status': 'ok', 'candidates': []}
    candidates = [nb for _, nb in rows if town_strip and town_strip[:2] in nb][:5]
    return None, {'status': 'town_not_matched', 'candidates': candidates}

def _posuto_search_pref_wide(pref, town):
    """都道府県内全体で町名を検索（合併済み市町村のフォールバック）

    完全一致または前方一致でユニークなものを返す。
    複数候補がある場合は None を返す（誤マッチ防止）。
    """
    conn = _get_posuto_conn()
    if conn is None or not town:
        return None
    def _norm(s):
        return s.replace('ヶ', 'ケ').replace('ヵ', 'カ').translate(_KANJI_VARIANTS)
    def _fmt(code):
        return f"{code[:3]}-{code[3:]}"
    town_n = _norm(town)
    town_strip = re.sub(r'[町村]$', '', town_n)
    rows = conn.execute(
        'SELECT code, city, neighborhood FROM postal_data WHERE prefecture=?',
        (pref,)).fetchall()
    # 完全一致を優先
    exact = [(c, ct, nb) for c, ct, nb in rows if _norm(nb) == town_n]
    if len(exact) == 1:
        return _fmt(exact[0][0])
    # 末尾町/村除き完全一致
    if town_strip and town_strip != town_n:
        strip_match = [(c, ct, nb) for c, ct, nb in rows
                       if re.sub(r'[町村]$', '', _norm(nb)) == town_strip]
        if len(strip_match) == 1:
            return _fmt(strip_match[0][0])
    return None

def _heartrails_town_search(pref, city, town, strict=False):
    """HeartRails getTowns API で郵便番号を検索

    strict=True のとき完全一致のみ（誤マッチ防止用）。
    戻り値: (postal | None, info)
      info = {'status': 'ok'|'city_not_found'|'town_not_matched', 'candidates': [...]}
    """
    locations = _get_towns(pref, city)
    if not locations:
        return None, {'status': 'city_not_found', 'candidates': []}
    def _kana_norm(s):
        return s.replace('ヶ', 'ケ').replace('ヵ', 'カ').translate(_KANJI_VARIANTS)
    town_clean = _kana_norm(re.sub(r'^[大小]?字', '', town))
    if not town_clean:
        return None, {'status': 'town_not_matched', 'candidates': []}
    all_towns = [loc.get('town', '') for loc in locations if loc.get('town')]
    # 末尾「町/村」を除いた正規化で一致判定（鼻毛石 ↔ 鼻毛石町 などをマッチ）
    town_stripped = re.sub(r'[町村]$', '', town_clean)
    for loc in locations:
        loc_town = _kana_norm(re.sub(r'^[大小]?字', '', loc.get('town', '')))
        loc_stripped = re.sub(r'[町村]$', '', loc_town)
        if strict:
            # 厳格モード: 「町/村」除き完全一致のみ
            if town_stripped and loc_stripped == town_stripped:
                p = loc['postal']
                return f"{p[:3]}-{p[3:]}", {'status': 'ok', 'candidates': []}
            continue
        # 通常マッチ: loc_town が town_clean の先頭部分（鼻毛石町←鼻毛石）
        if loc_town.startswith(town_clean):
            p = loc['postal']
            return f"{p[:3]}-{p[3:]}", {'status': 'ok', 'candidates': []}
        # 逆方向の包含マッチ: town_clean が loc_town で始まる
        # ただし loc_town が短い「○○町/村/区」型のときは誤マッチ防止のため除外
        # （沼田市の getTowns に「榛名町」だけがあり、群馬郡榛名町下里見が誤ヒットするのを回避）
        if town_clean.startswith(loc_town):
            if len(loc_town) <= 4 and loc_town.endswith(('町', '村', '区')):
                pass  # 誤マッチ防止: スキップ
            else:
                p = loc['postal']
                return f"{p[:3]}-{p[3:]}", {'status': 'ok', 'candidates': []}
        if town_stripped and loc_stripped == town_stripped:
            p = loc['postal']
            return f"{p[:3]}-{p[3:]}", {'status': 'ok', 'candidates': []}
    # 候補抽出: 共通文字数の多い順に上位5件
    scored = []
    town_chars = set(town_clean)
    for t in all_towns:
        t_norm = _kana_norm(re.sub(r'^[大小]?字', '', t))
        score = len(town_chars & set(t_norm))
        if score > 0:
            scored.append((score, t))
    scored.sort(key=lambda x: -x[0])
    candidates = [t for _, t in scored[:5]]
    if not candidates:
        candidates = all_towns[:5]
    return None, {'status': 'town_not_matched', 'candidates': candidates}

def lookup_postal_from_address(address):
    """住所から郵便番号を逆引き

    戻り値: (postal | None, fail_reason | None)
    """
    if not isinstance(address, str) or not address.strip():
        return None, "住所が空"
    pref_m = PREF_PATTERN.match(address)
    if not pref_m:
        return None, "都道府県を判定できず"
    pref = pref_m.group(1)
    rest = address[len(pref):]
    city, town = _extract_city_and_town(rest)
    if not city:
        return None, f"市区町村を抽出できず: 「{rest}」"
    if not town:
        return None, f"町名を抽出できず: 「{rest}」"

    # 最優先: posuto (日本郵便公式) で検索
    postal, info = _posuto_search(pref, city, town)
    if postal:
        return postal, None

    # posuto: 旧市 → 新市マッピング
    if (pref, city) in _OBSOLETE_CITY_MAP:
        for new_city in _OBSOLETE_CITY_MAP[(pref, city)]:
            postal_o, info_o = _posuto_search(pref, new_city, town)
            if postal_o:
                return postal_o, None

    # posuto: 県内ユニーク一致（合併済み市町村フォールバック）
    if '郡' in (city or ''):
        gun_m = re.match(r'^(.+)郡(.+)', city)
        if gun_m:
            cho_son = gun_m.group(2)
            # cho_son+town と town 単独を試す
            for t in [cho_son + town, town]:
                postal_u = _posuto_search_pref_wide(pref, t)
                if postal_u:
                    return postal_u, None
    else:
        postal_u = _posuto_search_pref_wide(pref, town)
        if postal_u:
            return postal_u, None

    # 続いて HeartRails で検索
    postal, info = _heartrails_town_search(pref, city, town)
    if postal:
        return postal, None

    # フォールバック: 廃止された旧市 → 新市（合併済み）
    # 例: 浦和市 → さいたま市浦和区、秋川市 → あきる野市
    if info.get('status') == 'city_not_found' and (pref, city) in _OBSOLETE_CITY_MAP:
        for new_city in _OBSOLETE_CITY_MAP[(pref, city)]:
            postal_o, info_o = _heartrails_town_search(pref, new_city, town)
            if postal_o:
                return postal_o, None
            if info_o.get('status') == 'town_not_matched' and info_o.get('candidates'):
                info = info_o

    # フォールバック: 大字なしで town に「○○町/村」が含まれる場合、
    # city に町/村まで含めて再検索（例: 小城市, 牛津町柿通瀬 → 小城市牛津町, 柿通瀬）
    sub_m = re.match(r'^([^\d一二三四五六七八九十]+(?:町|村))(.+)', town)
    if sub_m:
        city2 = city + sub_m.group(1)
        town2 = sub_m.group(2)
        postal2, info2 = _heartrails_town_search(pref, city2, town2)
        if postal2:
            return postal2, None
        if info2.get('status') == 'town_not_matched' and info2.get('candidates'):
            info = info2

    # フォールバック: 郡+旧町村の場合（合併済み自治体）
    # 例1: 神埼郡神埼町 → 神埼市+神埼町○○
    # 例2: 群馬郡榛名町 → 高崎市+榛名町○○（郡名と新市名が不一致）
    # 例3: 勢多郡粕川村 → 前橋市+粕川町○○（村→町に表記変化することも）
    if '郡' in city:
        gun_m = re.match(r'^(.+)郡(.+)', city)
        if gun_m:
            gun_base = gun_m.group(1)
            cho_son = gun_m.group(2)
            # cho_son の語尾を町/村両方試す
            cho_son_variants = [cho_son]
            if cho_son.endswith('村'):
                cho_son_variants.append(cho_son[:-1] + '町')
            elif cho_son.endswith('町'):
                cho_son_variants.append(cho_son[:-1] + '村')
            # 候補市: 郡名+市 を最優先、その後同県内の全市
            tried = set()
            candidate_cities = []
            primary = gun_base + '市'
            if primary in _CITY_PREF_CACHE:
                candidate_cities.append(primary)
            for c in _PREF_CITIES_CACHE.get(pref, []):
                if c not in candidate_cities:
                    candidate_cities.append(c)
            # 試す町名バリアント:
            #   通常モード: cho_son+town（合併後も旧町村名を保持: 群馬町福島）
            #   厳格モード: town 単独（合併後は旧町村名が消える: 子持村吹屋→渋川市吹屋）
            #     ※ 厳格モードは完全一致のみで誤マッチ防止
            non_strict_variants = [cs + town for cs in cho_son_variants]
            ku_fallback_info = None
            # フェーズ1: cho_son+town を通常マッチで全市試す
            for cand_city in candidate_cities:
                for tv in non_strict_variants:
                    if (pref, cand_city, tv, False) in tried:
                        continue
                    tried.add((pref, cand_city, tv, False))
                    postal3, info3 = _heartrails_town_search(pref, cand_city, tv)
                    if postal3:
                        return postal3, None
                    if info3.get('status') == 'town_not_matched' and info3.get('candidates') and ku_fallback_info is None:
                        ku_fallback_info = (cand_city, info3)
            # フェーズ2: town 単独を厳格マッチ
            # 候補市を「同郡内の現存町村 or 郡名+市」に絞って誤マッチを防ぐ
            # 例: 多野郡万場町 → 多野郡神流町（同郡）、利根郡月夜野町 → 利根郡みなかみ町（同郡）
            gun_full = gun_base + '郡'
            strict_candidates = []
            for c in _PREF_CITIES_CACHE.get(pref, []):
                if c == primary and c in _CITY_PREF_CACHE:
                    strict_candidates.append(c)
                elif c.startswith(gun_full):
                    strict_candidates.append(c)
            for cand_city in strict_candidates:
                if (pref, cand_city, town, True) in tried:
                    continue
                tried.add((pref, cand_city, town, True))
                postal3, info3 = _heartrails_town_search(pref, cand_city, town, strict=True)
                if postal3:
                    return postal3, None
            if ku_fallback_info:
                info = ku_fallback_info[1]

    # 失敗理由を組み立て
    if info.get('status') == 'city_not_found':
        return None, f"市区町村「{pref}{city}」がデータベースに見つからず"
    cands = info.get('candidates', [])
    if cands:
        return None, f"町名「{town}」が一致せず（候補: {', '.join(cands[:3])}）"
    return None, f"町名「{town}」が一致せず"

# ── 出力フォーマット（変換後FMT.xlsx 準拠の33列）────────────────────────
FMT_HEADERS = [
    '発注日','発送日','反響状況','反響日','オーナー名',
    '連名①','連名②','連名③','連名④','連名⑤',
    '郵便番号','オーナー住所','物件名①','物件名②','○○周辺',
    '物件住所①','物件住所②','地番','種別','備考',
    '時候の挨拶','日付','個人/法人','面積','差出人',
    '持ち分','階数','種別コード','予備1','予備2','予備3','予備4','予備5',
]

def _to_fmt_row(r):
    """整備済み行データ r を変換後FMTの列名→値の辞書に変換"""
    out = {h: '' for h in FMT_HEADERS}
    out['オーナー名'] = r.get('オーナー名', '') or ''
    for k in ['連名①','連名②','連名③','連名④','連名⑤']:
        out[k] = r.get(k, '') or ''
    out['郵便番号'] = r.get('郵便番号', '') or ''
    out['オーナー住所'] = r.get('オーナー住所', '') or ''
    out['物件名①'] = r.get('物件名', '') or ''
    out['物件住所①'] = r.get('物件住所', '') or ''
    out['地番'] = r.get('地番', '') or ''
    out['備考'] = r.get('備考', '') or ''
    return out

# ── 列マッピング ──────────────────────────────────────────────────────

def detect_columns(df):
    col_map = {k: None for k in
               ['オーナー名','連名①','連名②','連名③','連名④','連名⑤',
                '郵便番号','オーナー住所','物件名','物件住所','地番','備考']}
    alias_cols = []
    for col in df.columns:
        c = str(col)
        if any(k in c for k in ['所有者','オーナー','名義','氏名','代表者']) and col_map['オーナー名'] is None:
            col_map['オーナー名'] = col
        elif any(k in c for k in ['共有者','連名']):
            alias_cols.append(col)
        elif any(k in c for k in ['〒','zip','postal','郵便']):
            col_map['郵便番号'] = col
        elif any(k in c for k in ['居住地','オーナー住所','自宅','住所']) and '物件' not in c:
            col_map['オーナー住所'] = col
        elif any(k in c for k in ['物件名','店舗名','建物名','マンション名','物件名称']):
            col_map['物件名'] = col
        elif any(k in c for k in ['物件所在地','物件住所']):
            col_map['物件住所'] = col
        elif '地番' in c:
            col_map['地番'] = col
        elif '備考' in c:
            col_map['備考'] = col
    for i, ac in enumerate(alias_cols[:5]):
        col_map[f'連名{["①","②","③","④","⑤"][i]}'] = ac
    # 地番がヘッダー名で拾えなかった場合の補完:
    # 提供リストでは地番列が「物件住所」の直後に並ぶ。ヘッダーが空（Unnamed）の
    # 列はキーワード判定で捨てられるため、物件住所の右隣がヘッダー空なら地番とみなす。
    cols = list(df.columns)
    if col_map['地番'] is None and col_map['物件住所'] in cols:
        idx = cols.index(col_map['物件住所'])
        if idx + 1 < len(cols):
            nxt = cols[idx + 1]
            if str(nxt).startswith('Unnamed') and nxt not in col_map.values():
                col_map['地番'] = nxt
    return col_map

def get_val(row, col_map, key):
    col = col_map.get(key)
    if col is None:
        return ""
    v = row.get(col, "")
    if pd.isna(v):
        return ""
    # Excelが日付に自動変換した地番等（例: 「3-22」→2001-03-22）を文字列化
    if isinstance(v, (pd.Timestamp, datetime.datetime, datetime.date)):
        return v.strftime('%Y-%m-%d')
    return to_halfwidth(str(v)).strip()

# ── メイン処理 ────────────────────────────────────────────────────────

def process(file_bytes, progress_callback=None, sheet_name=0, manual_map=None):
    """
    file_bytes: bytes (Excelファイル)
    progress_callback: callable(message: str, progress: float) | None  進捗通知用
    sheet_name: int | str  読み込むシート（デフォルト=先頭シート）
    manual_map: dict {列名: 割当項目 | '使わない'} | None  空ヘッダー列等の手動上書き
    戻り値: (excel_bytes, summary_dict, error_list)
    """
    def notify(msg, progress=None):
        if progress_callback:
            progress_callback(msg, progress)

    df_raw = pd.read_excel(io.BytesIO(file_bytes), header=0, sheet_name=sheet_name)
    col_map = detect_columns(df_raw)

    # 手動マッピングで上書き（空ヘッダー列の割当をユーザーが確定したケース）
    if manual_map:
        for col, field in manual_map.items():
            if field in (None, '', '使わない'):
                # 自動割当されていた列を解除
                for k, v in list(col_map.items()):
                    if v == col:
                        col_map[k] = None
            elif field in col_map:
                col_map[field] = col

    # 空行除去：オーナー名・オーナー住所・郵便番号がすべて空の行はスキップ
    key_cols = [c for c in [col_map.get('オーナー名'), col_map.get('オーナー住所'), col_map.get('郵便番号')] if c]
    if key_cols:
        df_raw = df_raw[df_raw[key_cols].notna().any(axis=1)].reset_index(drop=True)

    logs, errors, raw_rows = [], [], []
    seen_keys = set()
    dup_count = addr_fill_count = postal_fill_count = merge_count = garbled_count = 0
    # 提供リストの各行が最終的にどうなったかの追跡 {orig_no: (区分, 詳細)}
    status_by_orig = {}

    # ── 行データ抽出 ──
    for idx, row in df_raw.iterrows():
        orig_no = idx + 2
        r = {
            'orig_no': orig_no,
            **{k: get_val(row, col_map, k)
               for k in ['オーナー名','連名①','連名②','連名③','連名④','連名⑤',
                         'オーナー住所','物件名','物件住所','地番','備考']},
            '郵便番号_raw': get_val(row, col_map, '郵便番号'),
            '郵便番号': '',
            '郵便番号失敗理由': '',
        }
        raw_rows.append(r)

    total = len(raw_rows)
    notify(f"住所・郵便番号を補完中... (全{total}件)", 0.05)

    # ── 郵便番号正規化 & 各種補完 ──
    for i, r in enumerate(raw_rows):
        if (i + 1) % 5 == 0 or i == total - 1:
            notify(f"住所・郵便番号を補完中... ({i+1}/{total}件)", 0.05 + 0.75 * (i + 1) / total)
        no = r['orig_no']
        postal_raw = r['郵便番号_raw']
        postal_norm = normalize_postal(postal_raw)

        if postal_norm and postal_norm != postal_raw:
            logs.append((no, f"郵便番号を正規化: 「{postal_raw}」→「{postal_norm}」"))
        r['郵便番号'] = postal_norm or ""

        # 都道府県補完（郵便番号優先 → 市区町村名から逆引き）
        if r['オーナー住所'] and not has_prefecture(r['オーナー住所']):
            orig = r['オーナー住所']
            pref = lookup_prefecture_from_postal(r['郵便番号'])
            method = "郵便番号から"
            if not pref:
                pref, _ = lookup_prefecture_from_city(orig)
                method = "市区町村名から"
                time.sleep(0.2)
            if pref:
                r['オーナー住所'] = pref + orig
                logs.append((no, f"都道府県を補完（{method}）: 「{orig}」→「{r['オーナー住所']}」"))
            else:
                logs.append((no, f"都道府県を特定できず: 「{orig}」"))

        # 郵便番号→住所補完（物件住所と一致する場合は禁止）
        if is_valid_postal(r['郵便番号']) and not r['オーナー住所']:
            filled = lookup_address_from_postal(r['郵便番号'])
            if filled:
                prop = r['物件住所']
                if prop and (prop.startswith(filled) or filled in prop):
                    logs.append((no, f"郵便番号({r['郵便番号']})の補完結果が物件住所と一致 → 物件の郵便番号と判断し補完を中止"))
                else:
                    logs.append((no, f"郵便番号からオーナー住所を補完: {r['郵便番号']}→「{filled}」"))
                    r['オーナー住所'] = filled
                    addr_fill_count += 1
            time.sleep(0.2)

        # 住所→郵便番号補完
        if r['オーナー住所'] and not is_valid_postal(r['郵便番号']):
            filled_postal, fail_reason = lookup_postal_from_address(r['オーナー住所'])
            if filled_postal:
                logs.append((no, f"オーナー住所から郵便番号を補完: 「{r['オーナー住所']}」→「{filled_postal}」"))
                r['郵便番号'] = filled_postal
                postal_fill_count += 1
            else:
                r['郵便番号失敗理由'] = fail_reason or ""
                logs.append((no, f"オーナー住所から郵便番号を逆引きできず: {fail_reason}"))
            time.sleep(0.5)  # API レート制限対策

        # フォールバック: 物件住所からの逆引き
        # オーナー住所が空 or 逆引き失敗時に、物件住所から郵便番号を取得
        # （オーナー＝物件所有者で自己居住している場合に有効）
        if not is_valid_postal(r['郵便番号']) and r['物件住所']:
            filled_postal, fail_reason2 = lookup_postal_from_address(r['物件住所'])
            if filled_postal:
                # オーナー住所がある場合、物件住所と一致するときのみ採用（誤適用防止）
                if r['オーナー住所']:
                    owner_norm = normalize_address_for_compare(r['オーナー住所'])
                    prop_norm = normalize_address_for_compare(r['物件住所'])
                    if owner_norm == prop_norm or (len(owner_norm) >= 10 and owner_norm[:10] == prop_norm[:10]):
                        r['郵便番号'] = filled_postal
                        r['郵便番号失敗理由'] = ''
                        postal_fill_count += 1
                        logs.append((no, f"物件住所から郵便番号を補完: 「{r['物件住所']}」→「{filled_postal}」（オーナー住所と一致）"))
                else:
                    # オーナー住所が空ならそのまま採用
                    r['郵便番号'] = filled_postal
                    r['郵便番号失敗理由'] = ''
                    postal_fill_count += 1
                    logs.append((no, f"物件住所から郵便番号を補完（オーナー住所空欄）: 「{r['物件住所']}」→「{filled_postal}」"))
            time.sleep(0.3)

    notify("重複削除・連名統合を処理中...", 0.82)

    # ── 重複削除 ──
    dedup_rows = []
    for r in raw_rows:
        key = (r['オーナー名'], normalize_address_for_compare(r['オーナー住所']), r['郵便番号'])
        if key in seen_keys and r['オーナー名']:
            logs.append((r['orig_no'], f"重複行として除外 (オーナー名: {r['オーナー名']}, 住所: {r['オーナー住所']})"))
            status_by_orig[r['orig_no']] = ('重複削除', 'オーナー名・住所・郵便番号が他行と重複')
            dup_count += 1
            continue
        seen_keys.add(key)
        dedup_rows.append(r)

    # ── 同一住所の連名統合（表記揺れ正規化後に比較）──
    addr_groups = {}
    for r in dedup_rows:
        if r['オーナー住所']:
            k = normalize_address_for_compare(r['オーナー住所'])
            addr_groups.setdefault(k, []).append(r)

    merged_ids = set()
    final_rows = []
    for r in dedup_rows:
        if id(r) in merged_ids:
            continue
        norm_key = normalize_address_for_compare(r['オーナー住所']) if r['オーナー住所'] else ""
        group = addr_groups.get(norm_key, [])
        if r['オーナー住所'] and len(group) > 1 and id(group[0]) == id(r):
            all_names = []
            for g in group:
                if g['オーナー名']:
                    all_names.append(g['オーナー名'])
                for k in ['連名①','連名②','連名③','連名④','連名⑤']:
                    if g[k]:
                        all_names.append(g[k])
                merged_ids.add(id(g))
                # 統合先（group[0]）以外は「連名統合」として記録
                if id(g) != id(group[0]):
                    status_by_orig[g['orig_no']] = (
                        '連名統合', f"{group[0]['orig_no']}行目に統合（{group[0]['オーナー名']}）")
            merged = dict(r)
            merged['オーナー名'] = all_names[0] if all_names else ""
            for i, ak in enumerate(['連名①','連名②','連名③','連名④','連名⑤']):
                merged[ak] = all_names[i+1] if i+1 < len(all_names) else ""
            final_rows.append(merged)
            nos = ','.join(str(g['orig_no']) for g in group)
            logs.append((nos, f"同一住所({r['オーナー住所']})の行を統合: {', '.join(all_names)}"))
            merge_count += 1
        elif r['オーナー住所'] and len(group) > 1:
            merged_ids.add(id(r))
        else:
            final_rows.append(r)

    notify("エラー判定・出力ファイル作成中...", 0.92)

    # ── エラー判定 ──
    ok_rows = []
    for r in final_rows:
        reasons = []
        for f in [r['オーナー名'], r['オーナー住所'], r['物件名'], r['物件住所']]:
            if is_garbled(f):
                reasons.append("文字化けの疑い")
                garbled_count += 1
                break
        if not r['オーナー名']:
            reasons.append("オーナー名が未入力")
        if not r['オーナー住所']:
            reasons.append("オーナー住所が未入力（補完不可）")
        if not is_valid_postal(r['郵便番号']):
            fail = r.get('郵便番号失敗理由', '')
            if fail:
                reasons.append(f"郵便番号を逆引きできず: {fail}")
            else:
                reasons.append("郵便番号が未入力または形式不正（逆引き不可）")
        if reasons:
            r['エラー理由'] = ' / '.join(reasons)
            errors.append(r)
            logs.append((r['orig_no'], f"エラーリストへ: {r['エラー理由']}"))
            status_by_orig[r['orig_no']] = ('エラー（送付不可）', r['エラー理由'])
        else:
            ok_rows.append(r)
            merged_note = '連名統合済み' if any(r.get(k) for k in
                ['連名①','連名②','連名③','連名④','連名⑤']) else ''
            status_by_orig[r['orig_no']] = ('採用', merged_note)

    # ── Excel出力 ──
    wb = openpyxl.Workbook()

    def style_header(cell, bg):
        cell.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        cell.fill = PatternFill('solid', start_color=bg)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    def set_col_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # 変換後FMT 33列の列幅（ヘッダー順に対応）
    fmt_widths = [
        10,10,10,10,15,  # 発注日,発送日,反響状況,反響日,オーナー名
        12,12,12,12,12,  # 連名①〜⑤
        13,35,18,18,12,  # 郵便番号,オーナー住所,物件名①,物件名②,○○周辺
        35,35,20,10,30,  # 物件住所①,物件住所②,地番,種別,備考
        12,12,10,10,12,  # 時候の挨拶,日付,個人/法人,面積,差出人
        10,8,12,8,8,8,8,8,  # 持ち分,階数,種別コード,予備1〜5
    ]

    # シート1: 整備済みリスト（変換後FMT.xlsx 準拠の33列）
    ws1 = wb.active
    ws1.title = "整備済みリスト"
    for ci, h in enumerate(FMT_HEADERS, 1):
        style_header(ws1.cell(row=1, column=ci, value=h), '1F4E79')
    for ri, r in enumerate(ok_rows, 2):
        fmt = _to_fmt_row(r)
        for ci, h in enumerate(FMT_HEADERS, 1):
            ws1.cell(row=ri, column=ci, value=fmt[h])
    set_col_widths(ws1, fmt_widths)
    ws1.row_dimensions[1].height = 20

    # シート2: エラーリスト（元行番号 + 変換後FMT33列 + エラー理由）
    ws2 = wb.create_sheet("エラーリスト")
    h2 = ['元行番号'] + FMT_HEADERS + ['エラー理由']
    for ci, h in enumerate(h2, 1):
        style_header(ws2.cell(row=1, column=ci, value=h), 'C00000')
    for ri, r in enumerate(errors, 2):
        fmt = _to_fmt_row(r)
        ws2.cell(row=ri, column=1, value=r.get('orig_no',''))
        for ci, h in enumerate(FMT_HEADERS, 2):
            ws2.cell(row=ri, column=ci, value=fmt[h])
        ws2.cell(row=ri, column=len(h2), value=r.get('エラー理由','') or '')
    set_col_widths(ws2, [10] + fmt_widths + [45])
    ws2.row_dimensions[1].height = 20

    # シート3: 提供リスト対応表（提供リストの各行が最終的にどうなったかを追跡）
    ws_t = wb.create_sheet("提供リスト対応表")
    orig_cols = [str(c) for c in df_raw.columns]
    h_t = orig_cols + ['整備状況', '処理詳細']
    for ci, h in enumerate(h_t, 1):
        style_header(ws_t.cell(row=1, column=ci, value=h), '7030A0')
    status_fills = {
        '採用': PatternFill('solid', start_color='E2EFDA'),
        '連名統合': PatternFill('solid', start_color='FFF2CC'),
        '重複削除': PatternFill('solid', start_color='D9D9D9'),
        'エラー（送付不可）': PatternFill('solid', start_color='FCE4D6'),
    }
    n_orig = len(orig_cols)
    for ri, (idx, row) in enumerate(df_raw.iterrows(), 2):
        orig_no = idx + 2
        for ci, c in enumerate(df_raw.columns, 1):
            v = row[c]
            ws_t.cell(row=ri, column=ci, value='' if pd.isna(v) else v)
        status, detail = status_by_orig.get(orig_no, ('未処理', ''))
        sc = ws_t.cell(row=ri, column=n_orig + 1, value=status)
        ws_t.cell(row=ri, column=n_orig + 2, value=detail)
        fill = status_fills.get(status)
        if fill:
            for ci in range(1, n_orig + 3):
                ws_t.cell(row=ri, column=ci).fill = fill
    set_col_widths(ws_t, [18] * n_orig + [16, 40])
    ws_t.row_dimensions[1].height = 20

    # シート4: 整備ログ
    ws3 = wb.create_sheet("整備ログ")
    for ci, h in enumerate(['行番号','処理内容'], 1):
        style_header(ws3.cell(row=1, column=ci, value=h), '375623')
    # 処理時のバージョン情報を1行目に記録（最新コードで処理されたか確認用）
    version = _get_app_version()
    from datetime import datetime as _dt
    ws3.cell(row=2, column=1, value='[INFO]')
    ws3.cell(row=2, column=2, value=f'処理バージョン: {version} ／ 処理日時: {_dt.now().strftime("%Y-%m-%d %H:%M:%S")}')
    for ri, (no, msg) in enumerate(logs, 3):
        ws3.cell(row=ri, column=1, value=str(no))
        ws3.cell(row=ri, column=2, value=msg)
    set_col_widths(ws3, [12, 90])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    excel_bytes = buf.read()

    summary = {
        '発注可能件数': len(ok_rows),
        'エラー件数': len(errors),
        '重複削除件数': dup_count,
        '住所補完件数': addr_fill_count,
        '郵便番号補完件数': postal_fill_count,
        '連名統合件数': merge_count,
        '文字化け検出件数': garbled_count,
    }
    error_list = [
        {'行番号': e['orig_no'], 'オーナー名': e.get('オーナー名',''), 'エラー理由': e.get('エラー理由','')}
        for e in errors
    ]
    return excel_bytes, summary, error_list
